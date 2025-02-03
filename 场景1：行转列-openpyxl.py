from openpyxl import load_workbook
import json

# pandas 是最常用且高效的解决方案，尤其是当你需要处理大型数据集时。
# openpyxl 提供了更多控制Excel格式的能力，适合需要细致控制的场景。
# numpy 更适用于数学计算密集型任务，但同样可以高效地处理数据转置。
# xlrd + xlwt 适用于老版本的Excel文件，但不再推荐用于现代的 .xlsx 文件。

# 假设这是你的JSON字符串
# json_str = '[{"pid":3,"pname":"颜色","vid":10061,"vname":"黑色"},{"pid":802,"pname":"规格","vid":-2,"vname":"40码"}]'


# df = pd.read_excel('ods_ic_item_sku_17218746.xlsx')
#
# df_transposed = df.transpose()
#
# # 将转置后的数据保存为新的Excel文件
# # df_transposed.to_excel('transposed_file.xlsx', index=False)  # 不保存索
#
# print(df)



# 加载Excel文件
wb = load_workbook('xlsx/ods_ic_item_sku_17218746.xlsx')
ws = wb['SheetJS']  # 选择你需要的工作表


# 将spec_value列的值更新为40码
# 遍历A列（列索引1是A列，'A'是列名）
for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=ws.max_row):
    for cell in row:
        # 修改spec_value列所有单元格的值
        data = json.loads(cell.value)
        newValue0 = data[0]['vname']
        newValue1 = data[1]['vname']
        cell.value = newValue0 + '-'+ newValue1

# 获取数据范围
data = []
for row in ws.iter_rows(values_only=True):
    data.append(row)

# 转置数据
transposed_data = list(zip(*data))

# 将转置后的数据写入新的工作表
ws_transposed = wb.create_sheet('Transposed')  # 创建一个新的工作表
for row_idx, row in enumerate(transposed_data, 1):
    for col_idx, value in enumerate(row, 1):
        ws_transposed.cell(row=row_idx, column=col_idx, value=value)

# 保存新的Excel文件
wb.save('xlsx/ods_ic_item_sku_17218746_openpyxl_output.xlsx')


# 遍历所有工作表并打印内容
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"工作表名: {sheet_name}")
    for row in sheet.iter_rows(values_only=True):
        print(row)  # 打印每一行的内容
    print("\n")  # 在工作表之间添加空行以便区分